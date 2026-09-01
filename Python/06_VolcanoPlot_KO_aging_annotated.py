import openpyxl
import json
import math
import matplotlib

matplotlib.use("Agg")

import matplotlib.pyplot as plt
import matplotlib.patches as mpatches


# ============================================================
# Annotated Volcano Plot
# KO: 14-month-old vs 6-month-old
# Phenotype-related gene categories
# Cadps2 excluded from visualization
# ============================================================

plt.rcParams["pdf.fonttype"] = 42
plt.rcParams["svg.fonttype"] = "none"


# ============================================================
# 1. Extract gene coordinates from Excel
#    and save them as JSON files
# ============================================================

wb = openpyxl.load_workbook(
    "Cadps2_7subgroups_classified_dirmatch.xlsx",
    data_only=True
)

sheets = [
    "Subgroup1_Aging-specific#1",
    "Subgroup2_Aging-specific#2",
    "Subgroup3_Aging_and_early_pancr",
    "Subgroup4_Aging_and_longstandin",
    "Subgroup5_Pancreatitis-specific",
    "Subgroup6_Longstanding_chronic_",
    "Subgroup7_Longstanding_chronic_"
]


# ------------------------------------------------------------
# Marker genes for each phenotype-related category
# ------------------------------------------------------------

markers = {

    "acinar": [
        "Amy1", "Amy2a5", "Amy2b",
        "Cela1", "Cela2a", "Cela3b",
        "Cpa1", "Cpa2", "Cpb1",
        "Prss1", "Prss2", "Prss3",
        "Ctrb1", "Ctrc",
        "Pnlip", "Clps",
        "Manf", "Tmed6", "Tmed11",
        "Cuzd1", "Serpini2",
        "Serpinb1a", "Retreg1",
        "Ptf1a", "Nr5a2", "Bhlha15", "Rbpjl"
    ],

    "secretion": [
        "Cadps", "Cadps2",
        "Rab27a", "Rab27b",
        "Syt7", "Syt1",
        "Snap25", "Stx1a", "Vamp2",
        "Unc13a", "Unc13b",
        "Rims1", "Rims2"
    ],

    "fibrosis_immune": [
        "Col1a1", "Col1a2", "Col3a1", "Col4a1",
        "Fn1", "Acta2", "Postn", "Pdgfrb",
        "Tgfb1", "Tgfb2", "Timp1",
        "Sparc", "Lox",
        "Cd3e", "Cd3d", "Cd3g",
        "Cd8a", "Cd4", "Cd68",
        "Ptprc", "Itgam",
        "Ccl2", "Cxcl1",
        "Il6", "Tnf", "Il1b", "Ccr2"
    ],

    "adm": [
        "Sox9", "Krt19", "Krt7",
        "Krt8", "Krt18",
        "Hnf1b", "Onecut1", "Muc1"
    ]
}


# ------------------------------------------------------------
# Classify genes according to phenotype-related markers
# ------------------------------------------------------------

def classify_gene(symbol):

    for category, genes in markers.items():

        if symbol in genes:
            return category

    return None


# ------------------------------------------------------------
# Collect gene information from all subgroup sheets
# ------------------------------------------------------------

all_genes = []

for sheet_name in sheets:

    ws = wb[sheet_name]

    header = [
        cell.value
        for cell in ws[1]
    ]

    column_index = {
        header_name: i
        for i, header_name in enumerate(header)
    }


    for row in ws.iter_rows(
        min_row=2,
        values_only=True
    ):

        symbol = row[column_index["Symbol"]]

        if symbol is None:
            continue


        log2fc = row[
            column_index["KO_aging_log2FC"]
        ]

        padj = row[
            column_index["KO_aging_padj"]
        ]

        category = classify_gene(symbol)


        all_genes.append({
            "symbol": symbol,
            "log2FC": log2fc,
            "padj": padj,
            "category": category,
            "subgroup": sheet_name
        })


# ============================================================
# 2. Prepare coordinates for the volcano plot
# ============================================================


# ------------------------------------------------------------
# Background genes
# Genes not assigned to any phenotype-related category
# ------------------------------------------------------------

background = [
    (
        round(gene["log2FC"], 2),
        round(
            min(
                -math.log10(
                    max(gene["padj"], 1e-300)
                ),
                60
            ),
            2
        )
    )

    for gene in all_genes

    if gene["category"] is None
]


# ------------------------------------------------------------
# Category-specific genes
# ------------------------------------------------------------

category_genes = {}

for gene in all_genes:

    if gene["category"]:

        category_genes.setdefault(
            gene["category"],
            []
        ).append({

            "x": round(
                gene["log2FC"],
                2
            ),

            "y": round(
                min(
                    -math.log10(
                        max(
                            gene["padj"],
                            1e-300
                        )
                    ),
                    60
                ),
                2
            ),

            "gene": gene["symbol"]
        })


# ------------------------------------------------------------
# Save coordinates as JSON
# ------------------------------------------------------------

with open(
    "ko_volcano_bg.json",
    "w"
) as f:

    json.dump(
        background,
        f
    )


with open(
    "ko_volcano_cats.json",
    "w"
) as f:

    json.dump(
        category_genes,
        f,
        ensure_ascii=False
    )


# ============================================================
# 3. Generate the annotated volcano plot
#    Cadps2 is excluded
# ============================================================


# ------------------------------------------------------------
# Category colors
# ------------------------------------------------------------

category_colors = {

    "acinar": "#2a78d6",

    "secretion": "#eb6834",

    "fibrosis_immune": "#e34948",

    "adm": "#1baf7a"
}


# ------------------------------------------------------------
# Category labels
# ------------------------------------------------------------

category_labels = {

    "acinar":
        "Acinar / secretory granule",

    "secretion":
        "Exocytosis machinery",

    "fibrosis_immune":
        "Fibrosis / immune infiltration",

    "adm":
        "ADM / ductal"
}


category_order = [
    "acinar",
    "secretion",
    "fibrosis_immune",
    "adm"
]


# ------------------------------------------------------------
# Create figure
# ------------------------------------------------------------

fig, ax = plt.subplots(
    figsize=(7.2, 6.0)
)


# ------------------------------------------------------------
# Plot background genes
# All KO-aging DEGs
# ------------------------------------------------------------

background_x = [
    min(point[0], 12)
    for point in background
]

background_y = [
    min(point[1], 60)
    for point in background
]


ax.scatter(
    background_x,
    background_y,
    s=8,
    color="#a3a299",
    alpha=0.35,
    linewidths=0,
    zorder=2
)


# ------------------------------------------------------------
# Label position adjustments
# ------------------------------------------------------------

label_offsets = {

    "Snap25": (-18, 5),

    "Rims2": (16, 5),

    "Serpinb1a": (-30, -2),

    "Cuzd1": (22, 2),

    "Itgam": (-14, 6),

    "Ccr2": (-12, -7)
}


# ============================================================
# Plot genes by phenotype-related category
# ============================================================

for category in category_order:

    points = [

        point

        for point in category_genes.get(
            category,
            []
        )

        if point["gene"] != "Cadps2"
    ]


    if not points:
        continue


    x_values = [
        min(point["x"], 12)
        for point in points
    ]

    y_values = [
        min(point["y"], 60)
        for point in points
    ]


    # Plot category-specific genes
    ax.scatter(
        x_values,
        y_values,
        s=55,
        color=category_colors[category],
        edgecolors="white",
        linewidths=0.7,
        zorder=4,
        label=category_labels[category]
    )


    # Add gene labels
    for point in points:

        x = min(point["x"], 12)

        y = min(point["y"], 60)


        dx, dy = label_offsets.get(
            point["gene"],
            (0, 5)
        )


        ax.annotate(

            point["gene"],

            (x, y),

            textcoords="offset points",

            xytext=(dx, dy),

            fontsize=7,

            ha="center",

            style="italic",

            color="#3a3936",

            zorder=5
        )


# ============================================================
# 4. Figure formatting
# ============================================================


ax.set_xlabel(
    "log2FC (KO 14m vs 6m)",
    fontsize=9
)

ax.set_ylabel(
    "-log10(padj)",
    fontsize=9
)


ax.set_title(
    "KO-aging DEGs by phenotype category (Cadps2 excluded)",
    fontsize=10.5,
    loc="left",
    pad=8
)


# Zero reference line
ax.axvline(
    0,
    color="#c3c2b7",
    linewidth=0.8,
    zorder=1
)


# Grid
ax.grid(
    color="#e1e0d9",
    linewidth=0.6,
    zorder=0
)


# Remove top and right borders
for spine in [
    "top",
    "right"
]:

    ax.spines[
        spine
    ].set_visible(False)


# Set left and bottom border colors
for spine in [
    "left",
    "bottom"
]:

    ax.spines[
        spine
    ].set_color("#c3c2b7")


ax.tick_params(
    length=0,
    labelsize=8
)


# ============================================================
# 5. Legend
# ============================================================


legend_handles = [

    mpatches.Patch(
        color="#a3a299",
        alpha=0.5,
        label="Other KO-aging DEGs"
    )
]


legend_handles += [

    mpatches.Patch(
        color=category_colors[category],
        label=category_labels[category]
    )

    for category in category_order
]


ax.legend(
    handles=legend_handles,
    loc="upper left",
    frameon=False,
    fontsize=7.3
)


# ============================================================
# 6. Save figures
# ============================================================


fig.savefig(
    "ko_volcano_no_cadps2.pdf",
    dpi=300,
    bbox_inches="tight"
)


fig.savefig(
    "ko_volcano_no_cadps2.png",
    dpi=300,
    bbox_inches="tight"
)


plt.close(fig)
